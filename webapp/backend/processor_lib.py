import re
import io
import os
import unicodedata
import pandas as pd
import config_manager
from exceptions import MissingColumnsError
import zipfile
import tarfile
import hashlib

def _normalize_text(s):
    """
    Lowercases, strips whitespace/BOM, and removes accents/diacritics so that
    column names and status/type values match consistently regardless of the
    source file's encoding or language (e.g. 'Número de Pedido' == 'numero de pedido',
    'Creación' == 'creacion'). Used everywhere we compare user-provided text
    against known keyword lists, so detection stays robust as new file variants
    (different encodings, locales, Binance UI text changes) show up.
    """
    if s is None:
        return ""
    s = str(s).replace('\ufeff', '').strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
    s = s.replace('_', ' ').replace('-', ' ')
    return s.strip()

# Statuses that mean a P2P/order row should NEVER be counted, regardless of
# language or exact wording used by the exchange. This is a BLACKLIST (rather
# than a whitelist of "completed" words) on purpose: new/unseen "completed"
# wordings (any language, any future Binance UI change) are included by
# default, while anything clearly not-final is excluded. This is what actually
# drives is_cancelled_transaction() and the P2P status filter below.
import difflib

_NEGATIVE_STATUS_KEYWORDS = (
    'cancel', 'rechaz', 'reject', 'fallid', 'fail', 'expir', 'incomplet',
    'incomplete', 'error', 'devuel', 'refund', 'pend', 'progres', 'progress',
    'esperando', 'waiting', 'apelaci', 'appeal', 'disput', 'en curso',
    'anulad', 'anulado', 'anulada', 'invalid', 'void', 'denied'
)

# Optional Imports for Extended Support
try:
    import py7zr
except ImportError:
    py7zr = None
    
try:
    import rarfile
except ImportError:
    rarfile = None

def get_usd_ars_rate(fecha=None):
    try:
        import db_manager
        year = 2025
        if fecha:
            if hasattr(fecha, 'year'):
                year = fecha.year
            else:
                s = str(fecha).strip()
                if len(s) >= 4 and s[:4].isdigit():
                    year = int(s[:4])
        settings = db_manager.get_tax_settings(year)
        return float(settings.get('usd_ars_exchange_rate', 1000.0) or 1000.0)
    except Exception:
        return 1000.0

KEY_CANDIDATES = {
    'fecha': ['fecha', 'date', 'datetime', 'fecha/hora', 'created_at', 'timestamp', 'time'],
    'tipo': ['tipo', 'type', 'operacion', 'operación', 'action', 'transaction type'],
    'moneda': ['moneda', 'currency', 'moneda_destino', 'asset', 'destino', 'symbol', 'coin'],
    'moneda_origen': ['moneda_origen', 'moneda origen', 'source_currency', 'origen', 'from_currency', 'source asset'],
    'monto': ['monto', 'amount', 'monto_destino', 'monto destino', 'quantity', 'total'],
    'monto_origen': ['monto_origen', 'monto origen', 'source_amount', 'monto_inicial', 'source quantity'],
    'precio': ['precio', 'price', 'rate', 'cotizacion', 'cotización', 'unit price'],
    'codigo_operacion': ['codigo_operacion', 'codigo de operacion', 'order_id', 'id_operacion', 'order id', 'txid', 'id']
}

def find_column_fuzzy(df, candidates, cutoff=0.8):
    """
    Finds the best matching column in a DataFrame for a list of candidate names.
    Employs a 3-pass resolution strategy:
      Pass 1: Exact normalized string equality.
      Pass 2: Substring & token-set containment.
      Pass 3: Fuzzy SequenceMatcher scoring (ratio >= cutoff).
    Returns the original column name from df.columns, or None if no match meets criteria.
    """
    if df is None or len(df.columns) == 0 or not candidates:
        return None

    df_cols = list(df.columns)
    norm_cols = [_normalize_text(c) for c in df_cols]
    norm_candidates = [_normalize_text(c) for c in candidates if c]

    # Pass 1: Exact normalized equality
    for cand in norm_candidates:
        for idx, norm_col in enumerate(norm_cols):
            if norm_col == cand:
                return df_cols[idx]

    # Pass 2: Substring / Token matching
    for cand in norm_candidates:
        if len(cand) < 3:
            continue
        cand_tokens = set(cand.split())
        for idx, norm_col in enumerate(norm_cols):
            if not norm_col:
                continue
            if cand in norm_col or (len(norm_col) >= 4 and norm_col in cand):
                return df_cols[idx]
            col_tokens = set(norm_col.split())
            if cand_tokens and cand_tokens.issubset(col_tokens):
                return df_cols[idx]

    # Pass 3: Fuzzy ratio matching via difflib
    best_match = None
    best_score = 0.0

    for cand in norm_candidates:
        if len(cand) < 3:
            continue
        for idx, norm_col in enumerate(norm_cols):
            if not norm_col:
                continue
            ratio = difflib.SequenceMatcher(None, cand, norm_col).ratio()
            if ratio > best_score and ratio >= cutoff:
                best_score = ratio
                best_match = df_cols[idx]

    return best_match

def validate_columns(df, exchange):
    """Checks if critical columns exist in the dataframe using fuzzy matching."""
    config = config_manager.load_config()
    if exchange not in config: return # Skip if no config
    
    required_map = config[exchange]['columns']
    missing = []
    
    for key, col_name in required_map.items():
        if not col_name: continue # Skip empty config
        cands = [col_name, key] + KEY_CANDIDATES.get(key, [])
        matched_col = find_column_fuzzy(df, cands)
        if matched_col is None:
            missing.append(f"{key} ('{col_name}')")
            
    if missing:
        available = [str(c)[:20] for c in df.columns]
        raise MissingColumnsError(exchange, missing, available)

# --- UTILS ---

def parse_date(date_str, exchange):
    """Parses date using configured format with resilient fallback."""
    fmt = config_manager.get_date_format(exchange)
    
    if fmt:
        try:
            parsed = pd.to_datetime(date_str, format=fmt)
            if not pd.isna(parsed):
                return parsed
        except Exception:
            try:
                # Resilient fallback if seconds or millisecond precision differs slightly
                parsed = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
                if not pd.isna(parsed):
                    return parsed
            except Exception:
                pass
            print(f"Date Parse Warning: '{date_str}' did not match exact format '{fmt}' for {exchange}. Using fallback.")
            
    # Default to Argentina (Day First)
    try:
        parsed = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
        if not pd.isna(parsed):
            return parsed
    except Exception:
        pass
    try:
        parsed = pd.to_datetime(date_str, errors='coerce')
        if not pd.isna(parsed):
            return parsed
    except Exception:
        pass
    # All parsers exhausted - return None to signal caller to skip this row
    print(f"Date Parse Error: '{date_str}' could not be parsed for {exchange}. Row will be skipped.")
    return None

def clean_decimal(val):
    if isinstance(val, (int, float)): return float(val)
    if pd.isna(val): return 0.0
    s = str(val).strip()
    if not s: return 0.0
    
    # Strip currency symbols and letters (e.g. '$ 50,00' -> '50,00')
    s = re.sub(r'[^\d.,\-+]', '', s).strip()
    if not s: return 0.0
    
    # Robust handling for 1.234,56 (Spanish) vs 1,234.56 (English)
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
             # Spanish: 1.234,56 -> Remove dots, replace comma
             s = s.replace('.', '').replace(',', '.')
        else:
             # English: 1,234.56 -> Remove commas
             s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
        
    try:
        return float(s)
    except Exception as e:
        # Final fallback: assume dots are thousand separators
        try:
            return float(s.replace('.', ''))
        except Exception as e2:
            print(f"Error parseando decimal '{val}': {e} / {e2}")
            return 0.0

def limpiar_numero_binance(valor):
    """Convierte '20USDT' -> 20.0 (float)"""
    if pd.isna(valor) or str(valor).strip() == '':
        return 0.0
    match = re.search(r"([\d\.]+)", str(valor))
    if match:
        return float(match.group(1))
    return 0.0
def is_cancelled_transaction(row):
    """
    Scans the row for any status, outcome, condition, or state column and
    returns True if the transaction is cancelled, rejected, failed, expired,
    pending, disputed, or voided.
    Protects valid completed transactions containing negative status keywords in positive
    or negated contexts (e.g., 'Completado sin error', 'Completed (refund: n/a)', 'Dispute Resolved').
    """
    status_keywords = {
        'estado', 'status', 'state', 'resultado', 'result',
        'condicion', 'condition', 'situacion', 'situación',
        'etapa', 'outcome', 'estatus', 'condición'
    }
    
    cols = row.keys() if hasattr(row, 'keys') else getattr(row, 'index', [])
    for col in cols:
        col_clean = _normalize_text(col)
        if any(kw in col_clean for kw in status_keywords):
            val_str = _normalize_text(row[col])
            if not val_str:
                continue
            
            # Sanitize positive / negated phrases where negative keywords appear in non-cancelled contexts
            val_cleaned = re.sub(
                r'\b(sin|no|without)\s+(error|errors|rechazo|rechazos|rechaz\w*|disputa|disputas|disput\w*|fallo|fallos|cancellation|cancellations|anulacion|anulaciones)\b|'
                r'\brefund\s*[:=\s]*n/?a\b|\brefund\s*[:=\s]*none\b|\bno\s+refund\b|\bwithout\s+refund\b|'
                r'\bdispute\s+resolved\b|\bdisputa\s+resuelta\b',
                '',
                val_str,
                flags=re.IGNORECASE
            )
            
            if any(kw in val_cleaned for kw in _NEGATIVE_STATUS_KEYWORDS):
                return True
    return False

def normalize_exchange_name(ex_name):
    if not ex_name or not str(ex_name).strip():
        return 'Otros'
    clean = str(ex_name).strip()
    clean_upper = clean.upper()
    if clean_upper in ('NONE', 'NAN', 'NULL', ''):
        return 'Otros'
    
    if clean in ('Bitso', 'Fiwind', 'Binance', 'Ripio', 'Lemon', 'OKX', 'Bybit', 'Bitget'):
        return clean
        
    if 'BINANCE P2P' in clean_upper or 'BINANCE_P2P' in clean_upper:
        return 'Binance P2P'
    elif 'BINANCE' in clean_upper:
        return 'Binance Spot' if ('SPOT' in clean_upper or clean_upper == 'BINANCE') else 'Binance P2P'
    elif 'BITSO' in clean_upper:
        return 'Bitso Alpha' if 'ALPHA' in clean_upper else 'Bitso'
    elif 'FIWIND' in clean_upper:
        return 'Fiwind'
    elif 'RIPIO TRADE' in clean_upper or 'RIPIO_TRADE' in clean_upper or 'RIPIO (CSV)' in clean_upper:
        return 'Ripio Trade'
    elif 'RIPIO CLASSIC' in clean_upper or 'RIPIO_CLASSIC' in clean_upper or clean_upper == 'RIPIO':
        return 'Ripio Classic'
    elif 'LEMON' in clean_upper:
        return 'Lemon Cash'
    elif 'BITGET' in clean_upper:
        return 'Bitget P2P'
    elif 'OKX' in clean_upper:
        return 'OKX'
    elif 'BYBIT' in clean_upper:
        return 'Bybit'
    
    return clean.title() if clean.isupper() or clean.islower() else clean

def create_transaction(fecha, exchange, tipo_op, moneda, m_compra, m_venta, cot_compra, cot_venta, m_ars, comentario="", unique_id=None):
    from models_v2 import compute_canonical_tx_hash
    exchange = normalize_exchange_name(exchange)
    ex_str = str(exchange).strip() if exchange is not None else ''
    if not ex_str or ex_str.lower() in ('nan', 'none', 'null', ''):
        exchange_name = 'Otros'
    else:
        exchange_name = ex_str

    if isinstance(fecha, pd.Timestamp):
        fecha_fmt = fecha.strftime('%Y-%m-%d %H:%M:%S')
    else:
        try:
            parsed_dt = pd.to_datetime(fecha, dayfirst=True, errors='coerce')
            if pd.isna(parsed_dt):
                parsed_dt = pd.to_datetime(fecha, errors='coerce')
            if not pd.isna(parsed_dt):
                fecha_fmt = parsed_dt.strftime('%Y-%m-%d %H:%M:%S')
            else:
                print(f"create_transaction: Could not parse date '{fecha}', row skipped.")
                return None
        except Exception:
            print(f"create_transaction: Exception parsing date '{fecha}', row skipped.")
            return None
    
    ref = str(unique_id) if unique_id is not None else str(comentario)
    tx_hash = compute_canonical_tx_hash(fecha_fmt, exchange_name, tipo_op, moneda, m_compra, m_venta, m_ars, ref)

    return {
        'Fecha': fecha_fmt,
        'Exchange': exchange_name,
        'Tipo de Operación': tipo_op,
        'Moneda': moneda,
        'Monto Compra (Cripto)': float(round(m_compra, 12)) if m_compra and m_compra > 0 else 0,
        'Monto Venta (Cripto)': float(round(m_venta, 12)) if m_venta and m_venta > 0 else 0,
        'Cotización Compra': float(round(cot_compra, 2)) if cot_compra > 0 else 0,
        'Cotización Venta': float(round(cot_venta, 2)) if cot_venta > 0 else 0,
        'Monto ARS': float(round(abs(m_ars), 2)),
        'Comentarios': comentario,
        'tx_hash': tx_hash
    }

# --- API NORMALIZATION ---

def process_api_trades(exchange_name, trades):
    """
    Normalizes a list of ccxt trades (or Ripio generic trades) 
    into the standard internal transaction format.
    """
    processed = []
    
    for t in trades:
        try:
            # Check if this is a CCXT trade (has 'symbol' and 'side')
            if 'symbol' in t and 'side' in t:
                # CCXT Format
                dt_str = t.get('datetime', '')
                if dt_str.endswith('Z'): dt_str = dt_str[:-1]
                try:
                    fecha = pd.to_datetime(dt_str)
                except Exception as e:
                    print(f"Error parseando fecha CCXT '{dt_str}': {e}")
                    fecha = pd.to_datetime('today')
                    
                symbol = str(t.get('symbol', ''))
                crypto = symbol.split('/')[0] if '/' in symbol else symbol
                side = str(t.get('side', '')).lower()
                
                amount_crypto = float(t.get('amount', 0))
                price = float(t.get('price', 0))
                cost = float(t.get('cost', amount_crypto * price)) # Fiat amount
                
                if side == 'buy':
                    processed.append(create_transaction(
                        fecha, exchange_name, 'Compra', crypto,
                        amount_crypto, 0, price, 0, cost, f"ID: {t.get('id', '')}", unique_id=t.get('id')
                    ))
                elif side == 'sell':
                    processed.append(create_transaction(
                        fecha, exchange_name, 'Venta', crypto,
                        0, amount_crypto, 0, price, cost, f"ID: {t.get('id', '')}", unique_id=t.get('id')
                    ))

            # Ripio Format handling (from /v4/user/trades endpoint)
            elif 'pair' in t and ('total_value' in t or 'cost' in t):
                dt_str = t.get('date', t.get('created_at', ''))
                try: fecha = pd.to_datetime(dt_str)
                except Exception as e:
                    print(f"Error parseando fecha Ripio '{dt_str}': {e}")
                    fecha = pd.to_datetime('today')
                
                crypto = str(t['pair']).split('_')[0]
                amount_crypto = float(t.get('amount', t.get('executed_amount', 0)))
                price = float(t.get('price', 0))
                cost = float(t.get('total_value', t.get('cost', t.get('total', 0))))
                
                # Try to extract the side. Ripio sometimes provides 'side' or 'type' for user trades
                side = str(t.get('side', t.get('type', ''))).lower()
                
                if 'buy' in side or 'compra' in side:
                    processed.append(create_transaction(
                        fecha, exchange_name, 'Compra', crypto,
                        amount_crypto, 0, price, 0, cost, f"ID: {t.get('id', '')}", unique_id=t.get('id')
                    ))
                elif 'sell' in side or 'venta' in side:
                    processed.append(create_transaction(
                        fecha, exchange_name, 'Venta', crypto,
                        0, amount_crypto, 0, price, cost, f"ID: {t.get('id', '')}", unique_id=t.get('id')
                    ))
                else: # Fallback if API changed format unpredictably
                    processed.append(create_transaction(
                        fecha, exchange_name, 'Operacion Ripio', crypto,
                        amount_crypto, 0, price, 0, cost, f"ID: {t.get('id', '')} Side: {side}", unique_id=t.get('id')
                    ))
        except Exception as e:
            print(f"Error normalizing API trade: {e}")
            continue
            
    return processed

# --- GENERIC ARCHIVE PROCESSOR ---

def process_archive(file_obj, filename, depth=0, state=None):
    """
    Recursively process files inside an archive (ZIP/TAR/7z/RAR).
    Returns aggregated processed transactions and raw samples.
    """
    if state is None:
        state = {"extracted_size": 0}
    if depth > 3:
        raise ValueError("Excedida la profundidad máxima de descompresión permitida (límite: 3).")
        
    MAX_EXTRACT_SIZE = 100 * 1024 * 1024 # 100 MB
    print(f"Processing Archive: {filename} (depth={depth})")
    processed = []
    raw_sample = [] # We might only keep samples from the first few files to avoid bloat
    
    try:
        # ZIP Support
        if zipfile.is_zipfile(file_obj):
            # print("  > Is ZIP")
            with zipfile.ZipFile(file_obj, 'r') as z:
                for member_name in z.namelist():
                    # Skip directories and hidden files (MacOS)
                    if member_name.endswith('/') or '__MACOSX' in member_name or member_name.startswith('.'):
                        continue
                        
                    with z.open(member_name) as f:
                        remaining = MAX_EXTRACT_SIZE - state["extracted_size"]
                        if remaining <= 0:
                            raise ValueError("El archivo comprimido supera el tamaño límite de extracción de 100MB.")
                        content = f.read(remaining + 1)
                        state["extracted_size"] += len(content)
                        if state["extracted_size"] > MAX_EXTRACT_SIZE:
                            raise ValueError("El archivo de extracción excede el límite permitido de 100MB.")
                            
                        sub_file = io.BytesIO(content)
                        
                        p, r = process_uploaded_file(sub_file, member_name, depth + 1, state)
                        processed.extend(p)
                        if len(raw_sample) < 5: raw_sample.extend(r)

        # 7z Support
        elif py7zr and (py7zr.is_7zfile(file_obj) or filename.endswith('.7z')):
             file_obj.seek(0)
             try:
                  import tempfile
                  import shutil
                  with tempfile.TemporaryDirectory() as temp_dir:
                      with py7zr.SevenZipFile(file_obj, mode='r') as z:
                          z.extractall(path=temp_dir)
                          
                      for root, dirs, files in os.walk(temp_dir):
                          for fname in files:
                              if fname.startswith('.') or '__MACOSX' in fname: continue
                              file_path = os.path.realpath(os.path.join(root, fname))
                              if not file_path.startswith(os.path.realpath(temp_dir)):
                                  continue
                              with open(file_path, 'rb') as f:
                                  remaining = MAX_EXTRACT_SIZE - state["extracted_size"]
                                  if remaining <= 0:
                                      raise ValueError("El archivo comprimido supera el tamaño límite de extracción de 100MB.")
                                  content = f.read(remaining + 1)
                                  state["extracted_size"] += len(content)
                                  if state["extracted_size"] > MAX_EXTRACT_SIZE:
                                      raise ValueError("El archivo de extracción excede el límite permitido de 100MB.")
                                  sub_file = io.BytesIO(content)
                                  p, r = process_uploaded_file(sub_file, fname, depth + 1, state)
                                  processed.extend(p)
                                  if len(raw_sample) < 5: raw_sample.extend(r)
             except Exception as e:
                  print(f"7z Extraction Error: {e}")
                  raise e

        # RAR Support
        elif rarfile and (rarfile.is_rarfile(file_obj) or filename.endswith('.rar')):
             # print("  > Is RAR")
             file_obj.seek(0)
             with rarfile.RarFile(file_obj) as rf:
                  for f in rf.infolist():
                      if f.isdir(): continue
                      with rf.open(f) as rfo:
                          remaining = MAX_EXTRACT_SIZE - state["extracted_size"]
                          if remaining <= 0:
                              raise ValueError("El archivo comprimido supera el tamaño límite de extracción de 100MB.")
                          content = rfo.read(remaining + 1)
                          state["extracted_size"] += len(content)
                          if state["extracted_size"] > MAX_EXTRACT_SIZE:
                              raise ValueError("El archivo de extracción excede el límite permitido de 100MB.")
                          sub_file = io.BytesIO(content)
                          p, r = process_uploaded_file(sub_file, f.filename, depth + 1, state)
                          processed.extend(p)
                          if len(raw_sample) < 5: raw_sample.extend(r)

        # TAR Support
        elif tarfile.is_tarfile(file_obj):
            # print("  > Is TAR")
            file_obj.seek(0)
            with tarfile.open(fileobj=file_obj, mode='r:*') as t:
                for member in t.getmembers():
                    if not member.isfile(): continue
                    
                    f = t.extractfile(member)
                    if f:
                        remaining = MAX_EXTRACT_SIZE - state["extracted_size"]
                        if remaining <= 0:
                            raise ValueError("El archivo comprimido supera el tamaño límite de extracción de 100MB.")
                        content = f.read(remaining + 1)
                        state["extracted_size"] += len(content)
                        if state["extracted_size"] > MAX_EXTRACT_SIZE:
                            raise ValueError("El archivo de extracción excede el límite permitido de 100MB.")
                        sub_file = io.BytesIO(content)
                        p, r = process_uploaded_file(sub_file, member.name, depth + 1, state)
                        processed.extend(p)
                        if len(raw_sample) < 5: raw_sample.extend(r)
                            
    except Exception as e:
        print(f"Error processing archive {filename}: {e}")
        err_msg = str(e).lower()
        if "rar" in err_msg or "unrar" in err_msg:
            raise ValueError("Para procesar archivos .RAR se requiere instalar UnRAR en Windows. Por favor extrae los archivos o conviértelos a .ZIP antes de subirlos.")
        raise e

    return processed, raw_sample

# --- PROCESSORS ---

def process_fiwind(file_obj, filename):
    processed = []
    raw_sample = []
    try:
        df = pd.read_excel(file_obj) if (filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls')) else pd.read_csv(file_obj)
        df.columns = [str(c).lower().strip() for c in df.columns]
        raw_sample = df.head(10).fillna('').astype(str).to_dict(orient='records')
        
        # Validation
        validate_columns(df, 'fiwind')

        c_date = find_column_fuzzy(df, [config_manager.get_column('fiwind', 'fecha'), 'fecha', 'date', 'datetime', 'fecha/hora', 'created_at', 'timestamp'])
        c_type = find_column_fuzzy(df, [config_manager.get_column('fiwind', 'tipo'), 'tipo', 'type', 'operacion', 'operación', 'action', 'transaction type'])
        c_curr = find_column_fuzzy(df, [config_manager.get_column('fiwind', 'moneda'), 'destination currency', 'destination_currency', 'moneda_destino', 'moneda destino', 'asset', 'destino', 'moneda'])
        c_curr_orig = find_column_fuzzy(df, [config_manager.get_column('fiwind', 'moneda_origen'), 'source currency', 'source_currency', 'moneda_origen', 'moneda origen', 'origen', 'from_currency'])
        c_amt = find_column_fuzzy(df, [config_manager.get_column('fiwind', 'monto'), 'destination amount', 'destination_amount', 'monto_destino', 'monto destino', 'amount', 'monto'])
        c_amt_orig = find_column_fuzzy(df, [config_manager.get_column('fiwind', 'monto_origen'), 'source amount', 'source_amount', 'monto_origen', 'monto origen', 'monto_inicial'])
        c_price = find_column_fuzzy(df, [config_manager.get_column('fiwind', 'precio'), 'precio', 'price', 'rate', 'cotizacion', 'cotización'])

        records = df.to_dict('records')
        for index, row in enumerate(records):
            if is_cancelled_transaction(row):
                continue
            try:
                fecha = parse_date(row.get(c_date), 'fiwind')
                tipo_raw = str(row.get(c_type, '')).upper()
                moneda_destino = str(row.get(c_curr, '')).upper()
                moneda_origen = str(row.get(c_curr_orig, '')).upper()
                monto_destino = clean_decimal(row.get(c_amt, 0))
                monto_origen = clean_decimal(row.get(c_amt_orig, 0))
                cotizacion = clean_decimal(row.get(c_price, 0))

                tipo_normalized = _normalize_text(tipo_raw)

                is_conversion = any(kw in tipo_normalized for kw in [
                    'conversion', 'convert', 'swap', 'exchange', 'trade', 'buy', 'sell', 'compra', 'venta'
                ])
                is_deposit = any(kw in tipo_normalized for kw in [
                    'deposito', 'deposit', 'ingreso', 'incoming', 'received', 'receive'
                ])
                is_withdrawal = any(kw in tipo_normalized for kw in [
                    'retiro', 'withdrawal', 'envio', 'transfer', 'sent', 'send', 'outgoing'
                ])

                if is_conversion:
                    if moneda_origen == 'ARS':
                        processed.append(create_transaction(
                            fecha, 'Fiwind', 'Compra', moneda_destino,
                            abs(monto_destino), 0, cotizacion, 0, abs(monto_origen), tipo_raw, unique_id=index
                        ))
                    elif moneda_destino == 'ARS':
                        processed.append(create_transaction(
                            fecha, 'Fiwind', 'Venta', moneda_origen,
                            0, abs(monto_origen), 0, cotizacion, abs(monto_destino), tipo_raw, unique_id=index
                        ))
                    else:
                        usd_ars_rate = get_usd_ars_rate(fecha)
                        if cotizacion > 10.0:
                            ars_value = abs(monto_origen) * cotizacion
                        elif cotizacion > 0:
                            ars_value = abs(monto_origen) * cotizacion * usd_ars_rate
                        elif moneda_origen in ('USD', 'USDT', 'USDC', 'DAI', 'BUSD'):
                            ars_value = abs(monto_origen) * usd_ars_rate
                        elif moneda_destino in ('USD', 'USDT', 'USDC', 'DAI', 'BUSD'):
                            ars_value = abs(monto_destino) * usd_ars_rate
                        else:
                            ref_prices = {
                                'BTC': 60000.0, 'ETH': 3000.0, 'SOL': 150.0, 'BNB': 600.0,
                                'AVAX': 30.0, 'LINK': 15.0, 'DOT': 7.0, 'XRP': 0.6, 'ADA': 0.5
                            }
                            if moneda_origen in ref_prices:
                                usd_val = abs(monto_origen) * ref_prices[moneda_origen]
                            elif moneda_destino in ref_prices:
                                usd_val = abs(monto_destino) * ref_prices[moneda_destino]
                            else:
                                usd_val = abs(monto_origen) * 100.0
                            ars_value = usd_val * usd_ars_rate

                        comment = f"[INTERCAMBIO: {moneda_origen}->{moneda_destino}]"
                        
                        cot_v = ars_value / abs(monto_origen) if monto_origen > 0 else 0.0
                        processed.append(create_transaction(
                            fecha, 'Fiwind', 'Venta', moneda_origen,
                            0, abs(monto_origen), 0, cot_v, ars_value, comment, unique_id=f"{index}_v"
                        ))
                        
                        cot_c = ars_value / abs(monto_destino) if monto_destino > 0 else 0.0
                        processed.append(create_transaction(
                            fecha, 'Fiwind', 'Compra', moneda_destino,
                            abs(monto_destino), 0, cot_c, 0, ars_value, comment, unique_id=f"{index}_c"
                        ))
                elif is_deposit:
                    target_coin = moneda_destino if (moneda_destino and moneda_destino != 'ARS') else (moneda_origen if moneda_origen else 'USDT')
                    qty = abs(monto_destino if monto_destino > 0 else monto_origen)
                    if target_coin == 'ARS':
                        m_ars = qty
                        cot_val = 1.0
                    elif target_coin in ('USD', 'USDT', 'USDC', 'DAI', 'BUSD'):
                        usd_ars_rate = get_usd_ars_rate(fecha)
                        rate = cotizacion if cotizacion > 10.0 else usd_ars_rate
                        m_ars = qty * rate
                        cot_val = rate
                    else:
                        if cotizacion > 10.0:
                            cot_val = cotizacion
                            m_ars = qty * cot_val
                        else:
                            usd_ars_rate = get_usd_ars_rate(fecha)
                            cot_val = usd_ars_rate
                            m_ars = qty * usd_ars_rate
                    processed.append(create_transaction(
                        fecha, 'Fiwind', 'Ingreso Cripto', target_coin,
                        qty, 0, cot_val, 0, m_ars, tipo_raw, unique_id=f"{index}_dep"
                    ))
                elif is_withdrawal:
                    target_coin = moneda_origen if (moneda_origen and moneda_origen != 'ARS') else (moneda_destino if moneda_destino else 'USDT')
                    qty = abs(monto_origen if monto_origen > 0 else monto_destino)
                    if target_coin == 'ARS':
                        m_ars = qty
                        cot_val = 1.0
                    elif target_coin in ('USD', 'USDT', 'USDC', 'DAI', 'BUSD'):
                        usd_ars_rate = get_usd_ars_rate(fecha)
                        rate = cotizacion if cotizacion > 10.0 else usd_ars_rate
                        m_ars = qty * rate
                        cot_val = rate
                    else:
                        if cotizacion > 10.0:
                            cot_val = cotizacion
                            m_ars = qty * cot_val
                        else:
                            usd_ars_rate = get_usd_ars_rate(fecha)
                            cot_val = usd_ars_rate
                            m_ars = qty * usd_ars_rate
                    processed.append(create_transaction(
                        fecha, 'Fiwind', 'Retiro Cripto', target_coin,
                        0, qty, 0, cot_val, m_ars, tipo_raw, unique_id=f"{index}_ret"
                    ))
            except Exception as e:
                print(f"Error procesando fila en Fiwind: {e}")
                continue
    except Exception as e:
        print(f"Error Fiwind: {e}")
        return [], []

    return processed, raw_sample

def process_ripio_trade(file_obj, filename):
    processed = []
    raw_sample = []
    try:
        df = pd.read_csv(file_obj)
        df.columns = [str(c).lower().strip() for c in df.columns]
        raw_sample = df.head(10).fillna('').astype(str).to_dict(orient='records')

        validate_columns(df, 'ripio_trade')

        col_fecha = find_column_fuzzy(df, [config_manager.get_column('ripio_trade', 'fecha'), 'fecha', 'date', 'created_at', 'timestamp'])
        col_monto = find_column_fuzzy(df, [config_manager.get_column('ripio_trade', 'monto'), 'monto', 'amount', 'cantidad', 'total'])
        col_moneda = find_column_fuzzy(df, [config_manager.get_column('ripio_trade', 'moneda'), 'moneda', 'currency', 'asset', 'symbol'])
        col_cod = find_column_fuzzy(df, [config_manager.get_column('ripio_trade', 'codigo_operacion'), 'codigo_operacion', 'codigo de operacion', 'order_id', 'id_operacion'])
        
        FIAT_STABLES = {'ARS', 'USD', 'USDT', 'USDC', 'DAI', 'BUSD', 'EUR', 'UXD'}

        if col_cod and col_cod in df.columns:
            grupos = df.groupby(col_cod)
            for nombre, grupo in grupos:
                if any(is_cancelled_transaction(r) for _, r in grupo.iterrows()):
                    continue
                
                row_ars = grupo[grupo[col_moneda].astype(str).str.upper() == 'ARS']
                if not row_ars.empty:
                    row_quote = row_ars
                    row_cripto = grupo[grupo[col_moneda].astype(str).str.upper() != 'ARS']
                    quote_currency = 'ARS'
                else:
                    row_stables = grupo[grupo[col_moneda].astype(str).str.upper().isin(FIAT_STABLES)]
                    if not row_stables.empty:
                        row_quote = row_stables.iloc[[0]]
                        row_cripto = grupo[~grupo.index.isin(row_quote.index)]
                        quote_currency = str(row_quote.iloc[0][col_moneda]).upper()
                    else:
                        row_quote = grupo.iloc[[0]]
                        row_cripto = grupo.iloc[[1]] if len(grupo) > 1 else grupo
                        quote_currency = str(row_quote.iloc[0][col_moneda]).upper()

                if not row_quote.empty and not row_cripto.empty:
                    try:
                        m_quote = clean_decimal(row_quote.iloc[0][col_monto])
                        m_cripto = clean_decimal(row_cripto.iloc[0][col_monto])
                        mon_cripto = str(row_cripto.iloc[0][col_moneda]).upper()
                        fecha_str = row_quote.iloc[0][col_fecha]
                        fecha = parse_date(fecha_str, 'ripio_trade')
                        usd_ars_rate = get_usd_ars_rate(fecha)
                        ref_prices = {
                            'BTC': 60000.0, 'ETH': 3000.0, 'SOL': 150.0, 'BNB': 600.0,
                            'AVAX': 30.0, 'LINK': 15.0, 'DOT': 7.0, 'XRP': 0.6, 'ADA': 0.5, 'EUR': 1.08
                        }

                        if quote_currency == 'ARS':
                            m_ars = abs(m_quote)
                        elif quote_currency in ('USD', 'USDT', 'USDC', 'DAI', 'BUSD'):
                            m_ars = abs(m_quote) * usd_ars_rate
                        elif quote_currency in ref_prices:
                            usd_val = abs(m_quote) * ref_prices[quote_currency]
                            m_ars = usd_val * usd_ars_rate
                        elif mon_cripto in ref_prices:
                            usd_val = abs(m_cripto) * ref_prices[mon_cripto]
                            m_ars = usd_val * usd_ars_rate
                        else:
                            m_ars = abs(m_quote) * 100.0 * usd_ars_rate

                        if m_quote < 0: # Compra
                            cot = m_ars / abs(m_cripto) if m_cripto != 0 else 0
                            processed.append(create_transaction(
                                fecha, 'Ripio Trade', 'Compra', mon_cripto,
                                abs(m_cripto), 0, cot, 0, m_ars, f"ID: {nombre}", unique_id=nombre
                            ))
                        else: # Venta
                            cot = m_ars / abs(m_cripto) if m_cripto != 0 else 0
                            processed.append(create_transaction(
                                fecha, 'Ripio Trade', 'Venta', mon_cripto,
                                0, abs(m_cripto), 0, cot, m_ars, f"ID: {nombre}", unique_id=nombre
                            ))
                    except Exception as e:
                        print(f"Error procesando grupo de Ripio Trade ({nombre}): {e}")
                        continue
    except Exception as e:
        print(f"Error Ripio: {e}")
        return [], []
    return processed, raw_sample

def process_bitso(file_obj, filename):
    processed = []
    raw_sample = []
    try:
        file_obj.seek(0)
        ext = filename.lower()
        if ext.endswith('.xlsx') or ext.endswith('.xls'):
            df = pd.read_excel(file_obj)
        else:
            df = None
            for enc in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
                for sep in [None, ',', ';', '\t']:
                    try:
                        file_obj.seek(0)
                        kwargs = {'encoding': enc}
                        if sep is None:
                            kwargs['sep'] = None
                            kwargs['engine'] = 'python'
                        else:
                            kwargs['sep'] = sep
                        df_csv = pd.read_csv(file_obj, **kwargs)
                        if df_csv is not None and not df_csv.empty and len(df_csv.columns) > 1:
                            df = df_csv
                            break
                    except Exception:
                        pass
                if df is not None:
                    break
            if df is None:
                file_obj.seek(0)
                try:
                    df = pd.read_excel(file_obj)
                except Exception:
                    pass

        if df is None or df.empty:
            return [], []

        # Standardize columns to lowercase and stripped for robust case-insensitivity
        df.columns = [str(c).lower().strip() for c in df.columns]
        
        raw_sample = df.head(10).fillna('').astype(str).to_dict(orient='records')
        
        # Config (lowercased and stripped to match df.columns)
        c_date = str(config_manager.get_column('bitso', 'datetime')).lower().strip()
        c_date_fallback = str(config_manager.get_column('bitso', 'date_fallback')).lower().strip()
        c_type = str(config_manager.get_column('bitso', 'type')).lower().strip()
        c_major = str(config_manager.get_column('bitso', 'major')).lower().strip()
        c_minor = str(config_manager.get_column('bitso', 'minor')).lower().strip()
        c_amt = str(config_manager.get_column('bitso', 'amount')).lower().strip()
        c_val = str(config_manager.get_column('bitso', 'value')).lower().strip()
        c_rate = str(config_manager.get_column('bitso', 'rate')).lower().strip()
        
        def _get_val(row, primary, candidates):
            if primary and primary in row.index and not pd.isna(row[primary]) and str(row[primary]).strip() != '':
                return row[primary]
            for cand in candidates:
                cand_clean = str(cand).lower().strip()
                for col in row.index:
                    if str(col).lower().strip() == cand_clean and not pd.isna(row[col]) and str(row[col]).strip() != '':
                        return row[col]
            return ''

        for index, row in df.iterrows():
            if is_cancelled_transaction(row):
                continue
            try:
                fecha_str = _get_val(row, c_date, [c_date_fallback, 'datetime', 'date', 'fecha', 'timestamp', 'created_at'])
                fecha = parse_date(fecha_str, 'bitso')
                
                tipo_raw = str(_get_val(row, c_type, ['type', 'tipo', 'side', 'operation'])).lower().strip()
                major = str(_get_val(row, c_major, ['major', 'asset', 'moneda', 'cripto', 'coin', 'symbol'])).upper().strip()
                minor = str(_get_val(row, c_minor, ['minor', 'fiat', 'quote', 'currency'])).upper().strip()
                cant_cripto = abs(clean_decimal(_get_val(row, c_amt, ['amount', 'cantidad', 'monto', 'qty'])))
                monto_fiat = abs(clean_decimal(_get_val(row, c_val, ['value', 'total', 'monto_ars', 'monto total', 'val'])))
                rate = clean_decimal(_get_val(row, c_rate, ['rate', 'price', 'precio', 'cotizacion', 'cotización']))

                if rate == 0 and cant_cripto > 0 and monto_fiat > 0:
                    rate = monto_fiat / cant_cripto
                elif monto_fiat == 0 and cant_cripto > 0 and rate > 0:
                    monto_fiat = cant_cripto * rate

                is_buy = 'buy' in tipo_raw or 'compra' in tipo_raw
                is_sell = 'sell' in tipo_raw or 'venta' in tipo_raw

                row_ex = _get_val(row, '', ['exchange', 'plataforma', 'entidad', 'broker'])
                row_ex_str = str(row_ex).strip() if row_ex and str(row_ex).strip().lower() not in ['', 'nan', 'none'] else ''
                ex_name = row_ex_str if row_ex_str else 'Bitso Alpha'

                if is_buy:
                    processed.append(create_transaction(
                        fecha, ex_name, 'Compra', major if major else 'USDT',
                        cant_cripto, 0, rate, 0, monto_fiat, 'Bitso Trade', unique_id=index
                    ))
                elif is_sell:
                    processed.append(create_transaction(
                        fecha, ex_name, 'Venta', major if major else 'USDT',
                        0, cant_cripto, 0, rate, monto_fiat, 'Bitso Trade', unique_id=index
                    ))
                elif cant_cripto > 0 or monto_fiat > 0:
                    processed.append(create_transaction(
                        fecha, ex_name, 'Compra', major if major else 'USDT',
                        cant_cripto, 0, rate, 0, monto_fiat, 'Bitso Trade', unique_id=index
                    ))
            except Exception as e:
                print(f"Error procesando fila en Bitso: {e}")
                continue
    except Exception as e:
         print(f"Error Bitso: {e}")
         return [], []
    return processed, raw_sample

def find_binance_column(df, candidates):
    # Normalize (accents/case/BOM-insensitive) so headers work regardless of
    # the file's source encoding or exact locale spelling (e.g. "Número",
    # "Numero", "NÚMERO" and a mis-decoded "N��mero" all resolve the same).
    cols = [_normalize_text(c) for c in df.columns]
    cand_clean_list = [_normalize_text(c) for c in candidates]
    # Pass 1: Exact match
    for cand_clean in cand_clean_list:
        for idx, col in enumerate(cols):
            if col == cand_clean:
                return df.columns[idx]
    # Pass 2: Substring match
    for cand_clean in cand_clean_list:
        for idx, col in enumerate(cols):
            if cand_clean in col or (len(col) >= 4 and col in cand_clean):
                return df.columns[idx]
    return None

def get_binance_dataframes(file_obj, filename):
    file_obj.seek(0)
    ext = filename.lower()
    raw_dfs = []
    
    if ext.endswith('.xlsx') or ext.endswith('.xls'):
        try:
            sheets_dict = pd.read_excel(file_obj, sheet_name=None)
            if isinstance(sheets_dict, dict):
                for sheet_name, df_sheet in sheets_dict.items():
                    if df_sheet is not None and not df_sheet.empty:
                        raw_dfs.append(df_sheet)
            elif isinstance(sheets_dict, pd.DataFrame):
                raw_dfs.append(sheets_dict)
        except Exception as e:
            file_obj.seek(0)
            try:
                raw_dfs.append(pd.read_excel(file_obj))
            except Exception:
                pass
    else:
        # Try CSV with different encodings and separators or Excel fallback
        for enc in ['utf-8', 'utf-8-sig', 'latin1', 'cp1252', 'utf-16']:
            for sep in [None, ',', ';', '\t']:
                file_obj.seek(0)
                try:
                    kwargs = {'encoding': enc}
                    if sep is None:
                        kwargs['sep'] = None
                        kwargs['engine'] = 'python'
                    else:
                        kwargs['sep'] = sep
                    df_csv = pd.read_csv(file_obj, **kwargs)
                    if df_csv is not None and not df_csv.empty and len(df_csv.columns) > 1:
                        raw_dfs.append(df_csv)
                        break
                except Exception:
                    pass
            if raw_dfs:
                break
        if not raw_dfs:
            file_obj.seek(0)
            try:
                sheets_dict = pd.read_excel(file_obj, sheet_name=None)
                if isinstance(sheets_dict, dict):
                    raw_dfs.extend(sheets_dict.values())
                elif isinstance(sheets_dict, pd.DataFrame):
                    raw_dfs.append(sheets_dict)
            except Exception:
                pass

    final_dfs = []
    for df in raw_dfs:
        if df is None or df.empty:
            continue
        cols_str = " ".join([str(c).lower() for c in df.columns])
        is_known = any(k in cols_str for k in ['order', 'orden', 'fiat', 'asset', 'activo', 'side', 'executed', 'ejecutado', 'created', 'creación', 'status', 'estado', 'tipo', 'precio', 'monto', 'cantidad'])
        if is_known:
            final_dfs.append(df)
        else:
            # Check if header row is shifted down by up to 5 rows
            found_skip = False
            for skip in range(1, min(6, len(df))):
                try:
                    row_vals = " ".join([str(v).lower() for v in df.iloc[skip-1].values])
                    if any(k in row_vals for k in ['order', 'orden', 'fiat', 'asset', 'activo', 'side', 'executed', 'ejecutado', 'created', 'creación', 'status', 'estado', 'tipo', 'precio', 'monto', 'cantidad']):
                        new_cols = [str(v).strip() for v in df.iloc[skip-1].values]
                        new_df = df.iloc[skip:].copy()
                        new_df.columns = new_cols
                        final_dfs.append(new_df)
                        found_skip = True
                        break
                except Exception:
                    pass
            if not found_skip:
                final_dfs.append(df)
                
    return final_dfs

def process_binance_csv(file_obj, filename):
    """Parses a single Binance CSV or Excel (.xlsx/.xls) file (Spot or P2P) in Spanish or English."""
    processed = []
    raw_sample = []
    try:
        dfs = get_binance_dataframes(file_obj, filename)
        if not dfs:
            return [], []
            
        for df in dfs:
            if df is None or df.empty:
                continue
                
            # Standardize columns to lowercase and stripped
            df.columns = [str(c).lower().strip() for c in df.columns]
            if len(raw_sample) < 10:
                raw_sample.extend(df.head(10).fillna('').astype(str).to_dict(orient='records'))

            # Find explicit exchange column if present
            ex_col = find_binance_column(df, ['exchange', 'plataforma', 'entidad', 'broker', 'origen', 'cuenta', 'exchange/plataforma', 'plataforma/exchange'])

            # Find columns for P2P (Bilingual Spanish / English + Account Statements)
            p2p_type = find_binance_column(df, ['tipo de orden', 'tipo de transacción', 'tipo de operación', 'tipo de operacion', 'order type', 'order_type', 'tipo', 'side', 'type', 'operación', 'operacion', 'operation', 'dirección', 'direccion'])
            p2p_fiat = find_binance_column(df, ['tipo de fiat', 'moneda fiduciaria', 'moneda fiat', 'fiat type', 'fiat_type', 'fiat currency', 'fiat', 'fiduciaria'])
            p2p_asset = find_binance_column(df, ['tipo de activo', 'activo cripto', 'criptomoneda', 'asset type', 'asset_type', 'activo', 'cripto', 'asset', 'crypto', 'moneda', 'coin', 'base asset'])
            p2p_status = find_binance_column(df, ['estado de la orden', 'order status', 'status', 'estado', 'state', 'situación', 'situacion'])
            p2p_created = find_binance_column(df, ['hora de creación', 'fecha de creación', 'fecha de la orden', 'created time', 'created_time', 'create time', 'fecha y hora', 'fecha/hora', 'fecha', 'hora', 'date', 'time', 'order time', 'utc_time', 'utc time', 'time(utc)'])
            p2p_qty = find_binance_column(df, ['monto cripto', 'cantidad cripto', 'quantity', 'cantidad', 'monto', 'amount', 'qty', 'crypto amount', 'change'])
            p2p_price = find_binance_column(df, ['precio unitario', 'unit price', 'price', 'precio', 'cotización', 'cotizacion', 'unit price (fiat)'])
            p2p_total = find_binance_column(df, ['precio total', 'monto total', 'total price', 'total_price', 'total (fiat)', 'monto (fiat)', 'importe total', 'total', 'monto fiat', 'monto total (fiat)'])

            # Find columns for Spot
            spot_side = find_binance_column(df, ['side', 'tipo', 'tipo de operación'])
            spot_exec = find_binance_column(df, ['executed', 'ejecutado', 'monto ejecutado'])
            spot_date = find_binance_column(df, ['date(utc)', 'date', 'fecha', 'time'])
            spot_pair = find_binance_column(df, ['pair', 'par', 'instrumento'])
            spot_price = find_binance_column(df, ['price', 'precio'])
            spot_amt = find_binance_column(df, ['amount', 'monto', 'total'])

            is_spot = (spot_side is not None and spot_exec is not None) or (spot_pair is not None and spot_side is not None)
            is_p2p = (p2p_fiat is not None or p2p_status is not None) and not is_spot

            if is_p2p:
                for index, row in df.iterrows():
                    if is_cancelled_transaction(row):
                        continue
                    try:
                        fecha_raw = row.get(p2p_created, '') if p2p_created else ''
                        fecha = parse_date(fecha_raw, 'binance_p2p')
                        
                        tipo_raw = str(row.get(p2p_type, '')).lower().strip() if p2p_type else ''
                        is_buy = tipo_raw in ('buy', 'compra', 'comprar') or 'buy' in tipo_raw or 'compra' in tipo_raw
                        
                        fiat = str(row.get(p2p_fiat, 'ARS')).upper().strip() if p2p_fiat else 'ARS'
                        asset = str(row.get(p2p_asset, 'USDT')).upper().strip() if p2p_asset else 'USDT'
                        
                        qty = clean_decimal(row.get(p2p_qty, 0)) if p2p_qty else 0.0
                        price = clean_decimal(row.get(p2p_price, 0)) if p2p_price else 0.0
                        total = clean_decimal(row.get(p2p_total, qty * price)) if p2p_total else (qty * price)
                        
                        row_ex = str(row.get(ex_col, '')).strip() if ex_col else ''
                        ex_name = row_ex if row_ex and row_ex.lower() not in ['', 'nan', 'none'] else 'Binance P2P'

                        is_ars = any(k in fiat for k in ['ARS', 'PESO', 'ARGENTIN', '$']) or not fiat or fiat == ''
                        if is_ars:
                            if is_buy:
                                processed.append(create_transaction(
                                    fecha, ex_name, 'Compra', asset,
                                    qty, 0, price, 0, total, 'P2P', unique_id=index
                                ))
                            else:
                                processed.append(create_transaction(
                                    fecha, ex_name, 'Venta', asset,
                                    0, qty, 0, price, total, 'P2P', unique_id=index
                                ))
                    except Exception as e:
                        print(f"Error procesando fila Binance P2P: {e}")
                        pass
                
            elif is_spot:
                for index, row in df.iterrows():
                    if is_cancelled_transaction(row):
                        continue
                    try:
                        fecha = parse_date(row.get(spot_date), 'binance_spot')
                        side = str(row.get(spot_side)).upper().strip()
                        is_buy = 'BUY' in side or 'COMPRA' in side
                        pair = str(row.get(spot_pair, '')).upper().strip()
                        precio = clean_decimal(row.get(spot_price, 0))
                        cant = limpiar_numero_binance(row.get(spot_exec, 0))
                        total = limpiar_numero_binance(row.get(spot_amt, cant * precio))
                        
                        row_ex = str(row.get(ex_col, '')).strip() if ex_col else ''
                        ex_name = row_ex if row_ex and row_ex.lower() not in ['', 'nan', 'none'] else 'Binance Spot'

                        if 'ARS' in pair:
                            crypto = pair.replace('ARS', '').replace('/', '').replace('_', '')
                            m_ars = total
                            cot_compra = precio if is_buy else 0.0
                            cot_venta = precio if not is_buy else 0.0
                            if is_buy:
                                processed.append(create_transaction(
                                    fecha, ex_name, 'Compra', crypto,
                                    cant, 0, cot_compra, 0, m_ars, 'Spot', unique_id=index
                                ))
                            else:
                                processed.append(create_transaction(
                                    fecha, ex_name, 'Venta', crypto,
                                    0, cant, 0, cot_venta, m_ars, 'Spot', unique_id=index
                                ))
                        else:
                            base_coin = pair
                            quote_coin = ""
                            for quote in ['USDT', 'USDC', 'BUSD', 'DAI', 'USD', 'EUR', 'BTC', 'ETH']:
                                if pair.endswith(quote):
                                    base_coin = pair[:-len(quote)]
                                    quote_coin = quote
                                    break
                                    
                            import db_manager
                            try:
                                settings = db_manager.get_tax_settings(fecha.year)
                                rate = settings.get('usd_ars_exchange_rate', 1000.0)
                            except Exception:
                                rate = 1000.0
                                
                            m_ars = total * rate
                            
                            if quote_coin:
                                comment_swap = f"[INTERCAMBIO: {quote_coin}->{base_coin}]" if is_buy else f"[INTERCAMBIO: {base_coin}->{quote_coin}]"
                                if is_buy:
                                    cot_c = m_ars / cant if cant > 0 else 0.0
                                    processed.append(create_transaction(
                                        fecha, ex_name, 'Compra', base_coin,
                                        cant, 0, cot_c, 0, m_ars, comment_swap, unique_id=f"{index}_c"
                                    ))
                                    cot_v = m_ars / total if total > 0 else 0.0
                                    processed.append(create_transaction(
                                        fecha, ex_name, 'Venta', quote_coin,
                                        0, total, 0, cot_v, m_ars, comment_swap, unique_id=f"{index}_v"
                                    ))
                                else:
                                    cot_v = m_ars / cant if cant > 0 else 0.0
                                    processed.append(create_transaction(
                                        fecha, ex_name, 'Venta', base_coin,
                                        0, cant, 0, cot_v, m_ars, comment_swap, unique_id=f"{index}_v"
                                    ))
                                    cot_c = m_ars / total if total > 0 else 0.0
                                    processed.append(create_transaction(
                                        fecha, ex_name, 'Compra', quote_coin,
                                        total, 0, cot_c, 0, m_ars, comment_swap, unique_id=f"{index}_c"
                                    ))
                            else:
                                cot_compra = (precio * rate) if is_buy else 0.0
                                cot_venta = (precio * rate) if not is_buy else 0.0
                                if is_buy:
                                    processed.append(create_transaction(
                                        fecha, ex_name, 'Compra', base_coin,
                                        cant, 0, cot_compra, 0, m_ars, 'Spot', unique_id=index
                                    ))
                                else:
                                    processed.append(create_transaction(
                                        fecha, ex_name, 'Venta', base_coin,
                                        0, cant, 0, cot_venta, m_ars, 'Spot', unique_id=index
                                    ))
                    except Exception as e:
                        print(f"Error procesando fila Binance Spot: {e}")
                        pass

    except MissingColumnsError:
        raise # Re-raise known error
    except Exception as e:
        print(f"Error Binance CSV: {e}")
        # [MODIFIED] Don't silence generic errors on Binance files during debugging
        if "binance" in filename.lower():
             raise Exception(f"Error critico procesando Binance: {str(e)}")
        return [], []
    return processed, raw_sample

def process_binance_zip(file_obj, filename):
    """Legacy wrapper for zip files, now delegated to generic archive processor."""
    return process_archive(file_obj, filename)

# --- DISPATCHER ---

def procesar_ripio_comun_txt(file_obj, filename):
    print(f"Procesando Ripio Común (TXT): {filename}")
    processed = []
    raw_sample = []
    
    try:
        content = file_obj.read()
        if isinstance(content, bytes):
            text = content.decode('utf-8')
        else:
            text = content
            
        lines = [line.strip() for line in text.replace('\r', '').split('\n') if line.strip()]
        
        # Sample for frontend
        for i in range(min(10, len(lines))):
            raw_sample.append({'Line': i+1, 'Content': lines[i]})

        i = 0
        while i < len(lines):
            line = lines[i]
            
            # Robust Date Regex (dd/mm/yyyy or yyyy-mm-dd)
            date_match = re.search(r"(\d{2}/\d{2}/\d{4})", line)
            
            if date_match:
                fecha_str = date_match.group(1)
                try:
                    fecha = pd.to_datetime(fecha_str, dayfirst=True)
                except Exception as e:
                    print(f"Error regex fecha TXT robusto '{fecha_str}': {e}")
                    i+=1; continue
                
                # Context Window: Look at previous 3 lines and next 10 lines
                # This helps identify the operation type even if formatting shifts specific lines
                prev_lines = lines[max(0, i-3):i]
                next_lines = lines[i+1:min(len(lines), i+12)]
                
                # Detect Operation Type from Context
                context_str = " ".join(prev_lines + [line] + next_lines).upper()
                
                if "CANCELAD" in context_str or "RECHAZAD" in context_str or "FALLID" in context_str or "ERROR" in context_str:
                    i += 1
                    continue
                
                tipo_bloque = "MOVIMIENTO"
                tipo_op = "MOVIMIENTO"
                
                if "COMPRA" in context_str:
                    tipo_op = "COMPRA"
                    tipo_bloque = "Compra Cripto"
                elif "VENTA" in context_str:
                    tipo_op = "VENTA"
                    tipo_bloque = "Venta Cripto"
                elif "CVU" in context_str or "TRANSFERENCIA" in context_str:
                    if "DEPÓSITO" in context_str or "RECIBIDA" in context_str:
                        tipo_op = "INGRESO FIAT"
                        tipo_bloque = "Depósito CVU"
                    else:
                        tipo_op = "RETIRO FIAT"
                        tipo_bloque = "Retiro CVU"
                elif "INTERCAMBIO" in context_str or "SWAP" in context_str:
                    tipo_op = "INTERCAMBIO"
                    tipo_bloque = "Swap Cripto"
                elif "INGRESO DE CRIPTO" in context_str:
                    tipo_op = "INGRESO CRIPTO"
                    tipo_bloque = "Depósito Cripto"
                elif "ENVÍO DE CRIPTO" in context_str or "ENVIO DE CRIPTO" in context_str:
                    tipo_op = "RETIRO CRIPTO"
                    tipo_bloque = "Retiro Cripto"
                
                # Money Extraction Logic
                # Look for patterns like "100.50 USDC" or "$ 5000 ARS" in the following lines
                montos = []
                monedas = []
                
                for sub in next_lines:
                    # Stop if we hit the next transaction (detected by date)
                    if re.search(r"\d{2}/\d{2}/\d{4}", sub):
                        break
                        
                    # Ignore Balance lines to prevent picking up "Saldo: 500000" as transaction amount
                    sub_upper = sub.upper()
                    if "SALDO" in sub_upper or "BALANCE" in sub_upper or "DISPONIBLE" in sub_upper:
                        continue
                        
                    # Regex for "1,234.50 CODE" or "$ 1.234,50"
                    # We normalize text removal of '$' first
                    clean_sub = sub.replace('$', '').strip()
                    
                    # Match Number + Space + Word (Coin)
                    # Use finditer to find ALL matches in the line (e.g. "0.00 ARS ... 25000.00 ARS")
                    matches = re.finditer(r"([\d\.,]+)\s+([A-Z]{3,5})", clean_sub)
                    
                    for match in matches:
                        val_str = match.group(1)
                        coin = match.group(2)
                        
                        # Anti-fragile float conversion
                        try:
                            val = clean_decimal(val_str)
                            if not isinstance(val, (int, float)):
                                val = float(val)
                        except Exception as e:
                            # print(f"DEBUG ERROR clean_decimal: {e}") 
                            val = 0.0
                            
                        if val > 0:
                            montos.append(val)
                            monedas.append(coin)
                            # print(f"DEBUG MATCH: {val} {coin}")
                        else:
                            pass
                            # print(f"DEBUG ZERO: {val_str} -> {val} {coin}")
                
                if len(montos) > 0:
                    exchange = "Ripio Trade"

                    # Logic mapping based on detected type
                    if tipo_op == "COMPRA":
                        # Usually logic: First amount ARS (negative/spend), Second amount Crypto (positive/buy)
                        # Or checking currencies
                        idx_crypto = -1
                        max_ars = 0.0
                        
                        for idx, m in enumerate(monedas):
                            if m == 'ARS':
                                if montos[idx] > max_ars:
                                    max_ars = montos[idx]
                            else: 
                                # First non-ARS is likely the crypto
                                if idx_crypto == -1: idx_crypto = idx
                        
                        if idx_crypto != -1:
                            m_crypto = montos[idx_crypto]
                            cotizacion = max_ars / m_crypto if m_crypto > 0 else 0
                            
                            processed.append(create_transaction(
                                fecha, exchange, 'Compra', monedas[idx_crypto],
                                m_crypto, 0, cotizacion, 0, max_ars, tipo_bloque, unique_id=i
                            ))
                            
                    elif tipo_op == "VENTA":
                        idx_crypto = -1
                        max_ars = 0.0
                        
                        for idx, m in enumerate(monedas):
                            if m == 'ARS': 
                                if montos[idx] > max_ars:
                                    max_ars = montos[idx]
                            else: 
                                if idx_crypto == -1: idx_crypto = idx
                            
                        if idx_crypto != -1:
                            m_crypto = montos[idx_crypto]
                            cotizacion = max_ars / m_crypto if m_crypto > 0 else 0
                            
                            processed.append(create_transaction(
                                fecha, exchange, 'Venta', monedas[idx_crypto],
                                0, m_crypto, 0, cotizacion, max_ars, tipo_bloque, unique_id=i
                            ))
                    # [MODIFIED] User requested ONLY Buy/Sell. Ignoring everything else.
                    else:
                        pass 


            i += 1
            
    except Exception as e:
        print(f"Error parseando Ripio TXT Robust: {e}")
        return [], []
        
    return processed, raw_sample

def process_uploaded_file(file_obj, filename, depth=0, state=None):
    filename = filename.lower()
    
    # 1. Prioritize Known Data Extensions
    # If it looks like a data file, treat it as such to avoid false positive "tar" detection on text files
    known_extensions = ('.csv', '.xls', '.xlsx', '.txt')
    if not filename.endswith(known_extensions):
        # 2. Check for Archive support if not a known data extension
        # Or if it explicitly has an archive extension
        is_archive_ext = filename.endswith(('.zip', '.tar', '.tar.gz', '.tgz', '.7z', '.rar'))
        
        is_zip = False
        is_tar = False
        is_7z = False
        is_rar = False
        
        # Check Signatures
        # ZIP
        try:
            file_obj.seek(0)
            if zipfile.is_zipfile(file_obj):
                # print("Detected ZIP signature")
                file_obj.seek(0)
                return process_archive(file_obj, filename, depth, state)
        except Exception as e:
            # Not a zip file or corrupted
            pass
        
        # 7z (Check before TAR because TAR can be aggressive with "truncated header" errors on binaries)
        if py7zr:
            try:
                file_obj.seek(0)
                if py7zr.is_7zfile(file_obj) or filename.endswith('.7z'):
                    file_obj.seek(0)
                    return process_archive(file_obj, filename, depth, state)
            except Exception as e:
                # Not a 7z file
                pass

        # RAR
        if rarfile:
            try:
                file_obj.seek(0)
                if rarfile.is_rarfile(file_obj):
                    file_obj.seek(0)
                    return process_archive(file_obj, filename, depth, state)
            except Exception as e:
                # Not a rar file
                pass

        # TAR (Check last as it is most prone to false positives/errors on random binary data)
        try:
            file_obj.seek(0)
            if tarfile.is_tarfile(file_obj):
                file_obj.seek(0)
                return process_archive(file_obj, filename, depth, state)
        except Exception as e:
            # Not a tar file
            pass

        # Extensions Fallback
        if is_archive_ext:
             # print(f"Detected archive by extension: {filename}")
             file_obj.seek(0)
             return process_archive(file_obj, filename, depth, state)
    
    file_obj.seek(0) # Reset pointer
    
    # 1. Inspect header columns for content-based detection.
    # Tries multiple encodings/separators (same idea as get_binance_dataframes)
    # instead of a single utf-8 attempt, so BOM/latin1/cp1252-encoded CSVs
    # (common from Windows/Excel exports) don't silently fall through to an
    # empty cols_set and rely solely on "binance" being in the filename.
    cols_set = set()
    ext = filename.lower()
    if ext.endswith('.xlsx') or ext.endswith('.xls'):
        try:
            file_obj.seek(0)
            preview_df = pd.read_excel(file_obj, nrows=2)
            if preview_df is not None and not preview_df.empty:
                cols_set = {_normalize_text(c) for c in preview_df.columns}
        except Exception:
            pass
    else:
        for enc in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
            for sep in [None, ',', ';', '\t']:
                try:
                    file_obj.seek(0)
                    kwargs = {'encoding': enc, 'nrows': 2, 'on_bad_lines': 'skip'}
                    if sep is None:
                        kwargs['sep'] = None
                        kwargs['engine'] = 'python'
                    else:
                        kwargs['sep'] = sep
                    preview_df = pd.read_csv(file_obj, **kwargs)
                    if preview_df is not None and not preview_df.empty and len(preview_df.columns) > 1:
                        cols_set = {_normalize_text(c) for c in preview_df.columns}
                        break
                except Exception:
                    pass
            if cols_set:
                break
    file_obj.seek(0)

    # 1. EXPLICIT EXCHANGE COLUMN OR CONSOLIDATED REPORT CHECK
    # Check if the file contains an explicit Exchange/plataforma/broker column or
    # consolidated presentation headers (Cotización Compra, Cotización Venta, etc.).
    # If so, route to process_multi_exchange_excel FIRST to preserve row exchange names.
    has_explicit_exchange_col = any(k in cols_set for k in [
        'exchange', 'plataforma', 'entidad', 'broker', 'origen', 'cuenta',
        'exchange/plataforma', 'plataforma/exchange'
    ])
    has_consolidated_cols = any(k in cols_set for k in [
        'cotizacion compra', 'cotizacion venta', 'monto compra (cripto)',
        'monto venta (cripto)', 'cotizacion_compra', 'cotizacion_venta'
    ])

    if has_explicit_exchange_col or has_consolidated_cols:
        try:
            file_obj.seek(0)
            multi_records, multi_sample = process_multi_exchange_excel(file_obj, filename)
            if multi_records:
                return multi_records, multi_sample
        except Exception as e:
            print(f"Multi-exchange explicit check error: {e}")

    # 2. BITSO FILE CHECK (CSV or XLSX, by filename or Bitso header columns)
    is_bitso_header = (
        ('major' in cols_set and 'minor' in cols_set) or
        ('rate' in cols_set and 'value' in cols_set and ('major' in cols_set or 'minor' in cols_set or 'type' in cols_set or 'amount' in cols_set or 'datetime' in cols_set or 'date' in cols_set)) or
        ('bitso' in cols_set)
    )
    is_bitso_file = "bitso" in filename.lower() or is_bitso_header

    if is_bitso_file:
        try:
            file_obj.seek(0)
            res = process_bitso(file_obj, filename)
            if res and res[0]:
                return res
        except MissingColumnsError:
            pass
        except Exception as e:
            print(f"Error in Bitso file processing: {e}")

    # 3. OTHER EXCHANGES BY FILENAME OR SPECIFIC HEADERS
    if "fiwind" in filename.lower() and (filename.endswith('.xls') or filename.endswith('.xlsx') or filename.endswith('.csv')):
        try:
            file_obj.seek(0)
            return process_fiwind(file_obj, filename)
        except MissingColumnsError:
            pass
    
    elif ("ripio trade" in filename.lower() or "ripio_trade" in filename.lower() or 'codigo_operacion' in cols_set or 'codigo de operacion' in cols_set) and not filename.endswith('.txt'):
        try:
            file_obj.seek(0)
            return process_ripio_trade(file_obj, filename)
        except MissingColumnsError:
            pass
        
    elif ("ripio" in filename.lower() or "correcciones" in filename.lower()) and filename.endswith('.txt'):
        try:
            file_obj.seek(0)
            return procesar_ripio_comun_txt(file_obj, filename)
        except Exception:
            pass

    # 4. BINANCE SPECIFIC FILE CHECK
    # Only route to Binance processor if the file is explicitly named Binance or has strong Binance P2P / Spot headers
    is_binance_named = "binance" in filename.lower()
    is_binance_p2p_headers = any(k in cols_set for k in [
        'tipo de orden', 'número de pedido', 'numero de pedido', 'tipo de fiat',
        'precio total', 'hora de creación', 'hora de creacion', 'tarifa de creador',
        'comisión de tomador', 'comision de tomador', 'order type', 'fiat type', 'asset type'
    ])
    is_binance_spot_headers = ('executed' in cols_set and 'side' in cols_set and ('pair' in cols_set or 'price' in cols_set))

    if is_binance_named or is_binance_p2p_headers or is_binance_spot_headers:
        try:
            file_obj.seek(0)
            res = process_binance_csv(file_obj, filename)
            if res and res[0]:
                return res
        except MissingColumnsError:
            pass
        except Exception as e:
            print(f"Error checking Binance P2P/Spot: {e}")

    # 5. DYNAMIC CUSTOM EXCHANGE DETECTION
    try:
        import db_manager
        all_ex = db_manager.get_all_exchanges()
        for ex in all_ex:
            ex_name_clean = ex['name'].lower()
            ex_id_clean = ex['id'].lower()
            if (ex_name_clean in filename.lower() or ex_id_clean in filename.lower()) and (filename.lower().endswith('.csv') or filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls')):
                file_obj.seek(0)
                df = pd.read_csv(file_obj) if filename.lower().endswith('.csv') else pd.read_excel(file_obj)
                res_df = process_dynamic_csv(df, ex['name'], ex.get('mapping', {}), ex.get('dateFormat', '%d/%m/%Y %H:%M:%S'))
                records = res_df.to_dict('records') if not res_df.empty else []
                sample = df.head(5).to_dict('records') if not df.empty else []
                if records:
                    return records, sample
    except Exception as e:
        print("Dynamic CSV check error:", e)

    # 6. MULTI-EXCHANGE & GENERIC CONSOLIDATED EXCEL/CSV PARSER FALLBACK
    try:
        file_obj.seek(0)
        multi_records, multi_sample = process_multi_exchange_excel(file_obj, filename)
        if multi_records:
            return multi_records, multi_sample
    except Exception as e:
        print("Multi-exchange Excel check error:", e)

    # 7. FINAL FALLBACK TO BINANCE CSV / EXCEL
    if (filename.lower().endswith('.csv') or filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls')):
        try:
            file_obj.seek(0)
            res = process_binance_csv(file_obj, filename)
            if res and res[0]:
                return res
        except MissingColumnsError:
            pass
        except Exception:
            pass

    return [], []

def generate_master_excel(transactions):
    if not transactions: return None
    
    output = io.BytesIO()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Consolidado"
    
    # Styles configuration
    STYLES = {
        'Binance Spot': {'fill': 'F3BA2F', 'font': '000000'},
        'Binance P2P': {'fill': 'F3BA2F', 'font': '000000'},
        'Binance': {'fill': 'F3BA2F', 'font': '000000'},
        'Bitso': {'fill': '29AC00', 'font': 'FFFFFF'},
        'Fiwind': {'fill': 'EFB41D', 'font': '000000'},
        'Ripio Trade': {'fill': '7A3CE3', 'font': 'FFFFFF'},
        'Ripio Classic': {'fill': '7A3CE3', 'font': 'FFFFFF'},
        'Ripio': {'fill': '7A3CE3', 'font': 'FFFFFF'},
        'OKX': {'fill': '111111', 'font': 'FFFFFF'},
        'Bybit': {'fill': 'F7A600', 'font': '000000'},
        'Bitget': {'fill': '00F0FF', 'font': '191F61'},
        'Generic': {'fill': '4B5563', 'font': 'FFFFFF'}
    }
    
    TYPE_STYLES = {
        'Compra': {'color': '2E7D32'}, # Dark Green
        'Venta': {'color': 'C62828'},  # Red
        'Otros': {'color': '1565C0'}   # Blue
    }
    
    cols_order = ['Fecha', 'Exchange', 'Tipo de Operación', 'Moneda', 
             'Monto Compra (Cripto)', 'Monto Venta (Cripto)', 
             'Cotización Compra', 'Cotización Venta', 'Monto ARS', 'Comentarios']
    
    # Convert list to DataFrame for easier grouping
    df_all = pd.DataFrame(transactions)
    
    # Ensure columns exist
    for c in cols_order:
        if c not in df_all.columns: df_all[c] = ''
            
    if 'Exchange' not in df_all.columns or df_all.empty:
        return None
        
    grouped_exchange = df_all.groupby('Exchange')
    current_row = 1
    
    for exchange_name, ex_group in grouped_exchange:
        # Get Exchange Style
        ex_style = STYLES.get(str(exchange_name).split(' ')[0], STYLES.get(exchange_name, STYLES['Generic']))
        if 'Binance' in str(exchange_name): ex_style = STYLES['Binance Spot']
        if 'Ripio' in str(exchange_name): ex_style = STYLES['Ripio Trade']
        if 'Bitso' in str(exchange_name): ex_style = STYLES['Bitso']
        if 'Fiwind' in str(exchange_name): ex_style = STYLES['Fiwind']
        if 'OKX' in str(exchange_name): ex_style = STYLES['OKX']
        if 'Bybit' in str(exchange_name): ex_style = STYLES['Bybit']
        if 'Bitget' in str(exchange_name): ex_style = STYLES['Bitget']
        
        # Write Title for Exchange Block
        ws.cell(row=current_row, column=1, value=f"REPORTE: {exchange_name.upper()}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color=ex_style['font'])
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color=ex_style['fill'], end_color=ex_style['fill'], fill_type="solid")
        current_row += 2
        
        # Categorize Types (Compra, Venta, and Otros)
        def categorize_tipo(t):
            t_upper = str(t).upper()
            if 'COMPRA' in t_upper: return 'Compra'
            if 'VENTA' in t_upper: return 'Venta'
            return 'Otros'
            
        ex_group = ex_group.copy()
        ex_group['TipoCategoria'] = ex_group['Tipo de Operación'].apply(categorize_tipo)
        
        # We want to iterate in a specific order: Compra, Venta, Otros
        for cat in ['Compra', 'Venta', 'Otros']:
            cat_group = ex_group[ex_group['TipoCategoria'] == cat]
            if cat_group.empty: continue
            
            # Sort chronologically
            cat_group = cat_group.sort_values(by='Fecha')
            
            # Write Sub-Title for Type (e.g. COMPRAS)
            cat_color = TYPE_STYLES.get(cat, TYPE_STYLES['Otros'])['color']
            ws.cell(row=current_row, column=1, value=f"{cat.upper()}S")
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color=cat_color)
            current_row += 1
            
            # Write Headers
            for col_idx, col_name in enumerate(cols_order, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='333333', end_color='333333', fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1
            
            start_data_row = current_row
            
            # Write Data
            for _, row_data in cat_group.iterrows():
                is_swap = '[INTERCAMBIO:' in str(row_data.get('Comentarios', ''))
                for col_idx, col_name in enumerate(cols_order, 1):
                    if col_name == 'Monto ARS' and is_swap:
                        val = ""
                    else:
                        val = row_data.get(col_name, '')
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal='center')
                    # Format numbers
                    if isinstance(val, (int, float)):
                        if 'Monto' in col_name or 'Cotización' in col_name:
                             cell.number_format = '#,##0.00'
                current_row += 1
                
            end_data_row = current_row - 1
            
            # Add Summary Row
            ws.cell(row=current_row, column=8, value="TOTAL M. ARS:")
            ws.cell(row=current_row, column=8).font = Font(bold=True)
            ws.cell(row=current_row, column=8).alignment = Alignment(horizontal='right')
            
            # Dynamic SUBTOTAL Formula for 'Monto ARS' (Column 9 -> I) (109 = SUM excluding hidden/filtered rows)
            sum_cell = ws.cell(row=current_row, column=9)
            sum_cell.value = f"=SUBTOTAL(109, I{start_data_row}:I{end_data_row})"
            sum_cell.font = Font(bold=True, color=cat_color)
            sum_cell.number_format = '#,##0.00'
            sum_cell.alignment = Alignment(horizontal='center')
            
            # Add thin border to summary
            border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
            sum_cell.border = border
            
            current_row += 3 # Space between Compras and Ventas
            
        current_row += 5 # Space between Exchanges

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                # Don't use formulas for length calc
                if str(cell.value).startswith('='): continue 
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                pass
        adjusted_width = min(max_length + 2, 50) # Cap width to 50
        ws.column_dimensions[column].width = adjusted_width

    wb.save(output)
    output.seek(0)
    return output

# Alias for backward compatibility if needed, though we should update app.py
def generate_excel_bytes(transactions):
    return generate_master_excel(transactions)

def process_dynamic_csv(df, exchange_name, mapping, date_format="%d/%m/%Y %H:%M:%S"):
    """Generic CSV parser driven by user-configured mapping."""
    if df is None or df.empty:
        return pd.DataFrame()

    # Standardize columns to lowercase and stripped for robust case-insensitivity
    df = df.rename(columns=lambda c: str(c).lower().strip())
    
    # Standardize mapping to match lowercased columns
    clean_mapping = {k: str(v).lower().strip() for k, v in mapping.items() if v}

    records = []
    for idx, row in df.iterrows():
        if is_cancelled_transaction(row):
            continue
        raw_date = row.get(clean_mapping.get("fecha", ""), "")
        raw_tipo = row.get(clean_mapping.get("tipo_operacion", ""), "COMPRA")
        raw_moneda = row.get(clean_mapping.get("moneda", ""), "USDT")
        raw_compra = row.get(clean_mapping.get("monto_compra_cripto", ""), 0)
        raw_venta = row.get(clean_mapping.get("monto_venta_cripto", ""), 0)
        raw_cot_compra = row.get(clean_mapping.get("cotizacion_compra", ""), 0)
        raw_cot_venta = row.get(clean_mapping.get("cotizacion_venta", ""), 0)
        raw_ars = row.get(clean_mapping.get("monto_ars", ""), 0)
        raw_notes = row.get(clean_mapping.get("comentarios", ""), "")

        try:
            parsed_date = pd.to_datetime(raw_date, format=date_format) if date_format else pd.to_datetime(raw_date, dayfirst=True)
            formatted_date = parsed_date.strftime('%Y-%m-%d %H:%M:%S')
        except Exception:
            formatted_date = str(raw_date)

        def safe_float(val):
            try:
                val_str = str(val).replace('$', '').replace(',', '.').strip()
                return float(val_str)
            except Exception:
                return 0.0

        m_compra = safe_float(raw_compra)
        m_venta = safe_float(raw_venta)
        c_compra = safe_float(raw_cot_compra)
        c_venta = safe_float(raw_cot_venta)
        m_ars = safe_float(raw_ars)
        from models_v2 import compute_canonical_tx_hash
        ref = str(raw_notes) if raw_notes else str(idx)
        tx_hash = compute_canonical_tx_hash(formatted_date, exchange_name, raw_tipo, raw_moneda, m_compra, m_venta, m_ars, ref)

        records.append({
            "tx_hash": tx_hash,
            "fecha": formatted_date,
            "exchange": exchange_name,
            "tipo_operacion": str(raw_tipo).upper(),
            "moneda": str(raw_moneda).upper(),
            "monto_compra_cripto": m_compra,
            "monto_venta_cripto": m_venta,
            "cotizacion_compra": c_compra,
            "cotizacion_venta": c_venta,
            "monto_ars": m_ars,
            "comentarios": str(raw_notes)
        })

    return pd.DataFrame(records)


def process_multi_exchange_excel(file_obj, filename):
    """Parses a multi-exchange manual Excel/CSV file using a single-pass state machine.
    Returns (records, sample_records).
    """
    import openpyxl
    import datetime

    file_obj.seek(0)
    ext = filename.lower()
    rows = []

    try:
        if ext.endswith('.xlsx') or ext.endswith('.xls'):
            wb = openpyxl.load_workbook(file_obj, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        else:
            import csv
            for enc in ['utf-8-sig', 'utf-8', 'latin1', 'cp1252']:
                file_obj.seek(0)
                try:
                    content = file_obj.read()
                    if isinstance(content, bytes):
                        content = content.decode(enc)
                    dialect = None
                    try:
                        dialect = csv.Sniffer().sniff(content[:2048])
                    except Exception:
                        pass
                    sep = dialect.delimiter if dialect else ','
                    reader = csv.reader(io.StringIO(content), delimiter=sep)
                    rows = [r for r in reader]
                    if rows:
                        break
                except Exception:
                    continue
    except Exception as e:
        print(f"Error reading multi-exchange file: {e}")
        return [], []

    if not rows:
        return [], []

    current_exchange = 'Otros'
    col_map = None
    records = []

    for idx, row in enumerate(rows, start=1):
        try:
            if not row or not any(c is not None and str(c).strip() != '' for c in row):
                continue

            row_strs = [str(c).strip() if c is not None else '' for c in row]
            first_cell = row_strs[0] if row_strs else ''

            banner_match = re.search(r'REPORTE:\s*(.+)', first_cell, re.IGNORECASE)
            if not banner_match:
                for cell in row_strs:
                    bm = re.search(r'REPORTE:\s*(.+)', cell, re.IGNORECASE)
                    if bm:
                        banner_match = bm
                        break

            if banner_match:
                current_exchange = banner_match.group(1).strip()
                continue

            normalized_row = [_normalize_text(c) for c in row_strs]
            if any(h in normalized_row for h in ['fecha', 'date', 'timestamp', 'fecha/hora', 'datetime']):
                col_map = {}
                for col_i, norm_val in enumerate(normalized_row):
                    if not norm_val:
                        continue
                    if norm_val in ['fecha', 'date', 'timestamp', 'fecha/hora', 'datetime', 'created_at', 'hora'] and 'date' not in col_map:
                        col_map['date'] = col_i
                    elif norm_val in ['exchange', 'plataforma', 'entidad', 'broker', 'origen', 'cuenta', 'exchange/plataforma', 'plataforma/exchange'] and 'exchange' not in col_map:
                        col_map['exchange'] = col_i
                    elif norm_val in ['tipo', 'tipo de operacion', 'tipo de operación', 'type', 'operacion', 'operación', 'action', 'side'] and 'type' not in col_map:
                        col_map['type'] = col_i
                    elif norm_val in ['moneda', 'asset', 'cripto', 'ticker', 'coin', 'symbol', 'currency'] and 'asset' not in col_map:
                        col_map['asset'] = col_i
                    elif norm_val in ['monto compra (cripto)', 'monto compra', 'cantidad compra', 'monto_compra_cripto', 'buy_amount', 'compra'] and 'buy' not in col_map:
                        col_map['buy'] = col_i
                    elif norm_val in ['monto venta (cripto)', 'monto venta', 'cantidad venta', 'monto_venta_cripto', 'sell_amount', 'venta'] and 'sell' not in col_map:
                        col_map['sell'] = col_i
                    elif norm_val in ['cantidad', 'monto', 'amount', 'quantity', 'volume'] and 'amount' not in col_map:
                        col_map['amount'] = col_i
                    elif norm_val in ['monto ars', 'monto_ars', 'total ars', 'monto usd', 'total', 'total fiat', 'monto fiat', 'fiat', 'ars'] and 'fiat' not in col_map:
                        col_map['fiat'] = col_i
                    elif norm_val in ['cotización compra', 'cotizacion compra', 'cotización_compra', 'cotizacion_compra', 'precio compra'] and 'cot_buy' not in col_map:
                        col_map['cot_buy'] = col_i
                    elif norm_val in ['cotización venta', 'cotizacion venta', 'cotización_venta', 'cotizacion_venta', 'precio venta'] and 'cot_sell' not in col_map:
                        col_map['cot_sell'] = col_i
                    elif norm_val in ['precio', 'price', 'cotizacion', 'cotización', 'rate'] and 'price' not in col_map:
                        col_map['price'] = col_i
                    elif norm_val in ['comentarios', 'comentario', 'notes', 'memo', 'detalle', 'observaciones', 'referencia'] and 'notes' not in col_map:
                        col_map['notes'] = col_i
                continue

            if not col_map:
                continue

            if len([c for c in row_strs if c]) == 1 and row_strs[0].upper() in ['COMPRAS', 'VENTAS', 'DEPOSITOS', 'RETIROS', 'TRANSACCIONES']:
                continue

            date_idx = col_map.get('date')
            if date_idx is None or date_idx >= len(row):
                continue

            raw_date = row[date_idx]
            if raw_date is None or str(raw_date).strip() == '' or str(raw_date).strip().lower() in ['fecha', 'date', 'nan', 'none']:
                continue

            if any('total' in str(c).lower() for c in row_strs if c):
                continue

            dt = parse_date(raw_date, current_exchange)

            ex_idx = col_map.get('exchange')
            raw_ex_val = str(row[ex_idx]).strip() if ex_idx is not None and ex_idx < len(row) and row[ex_idx] is not None else ''
            if raw_ex_val and raw_ex_val.lower() not in ('', 'nan', 'none', 'null'):
                row_ex = raw_ex_val
            else:
                row_ex = current_exchange if current_exchange and current_exchange.lower() not in ('', 'nan', 'none', 'null') else 'Otros'

            type_idx = col_map.get('type')
            raw_tipo_val = str(row[type_idx]).strip() if type_idx is not None and type_idx < len(row) and row[type_idx] is not None and str(row[type_idx]).strip() not in ['', 'nan', 'None'] else 'COMPRA'
            if 'COMPRA' in raw_tipo_val.upper() or 'BUY' in raw_tipo_val.upper():
                tipo_op = 'Compra'
            elif 'VENTA' in raw_tipo_val.upper() or 'SELL' in raw_tipo_val.upper():
                tipo_op = 'Venta'
            else:
                tipo_op = raw_tipo_val.title() if (raw_tipo_val.isupper() or raw_tipo_val.islower()) else raw_tipo_val

            asset_idx = col_map.get('asset')
            raw_moneda = str(row[asset_idx]).strip().upper() if asset_idx is not None and asset_idx < len(row) and row[asset_idx] is not None and str(row[asset_idx]).strip() not in ['', 'nan', 'None'] else 'USDT'

            buy_idx = col_map.get('buy')
            sell_idx = col_map.get('sell')
            amount_idx = col_map.get('amount')

            m_compra = clean_decimal(row[buy_idx]) if buy_idx is not None and buy_idx < len(row) else 0.0
            m_venta = clean_decimal(row[sell_idx]) if sell_idx is not None and sell_idx < len(row) else 0.0

            if m_compra == 0.0 and m_venta == 0.0 and amount_idx is not None and amount_idx < len(row):
                amt = clean_decimal(row[amount_idx])
                if any(kw in raw_tipo_val.upper() for kw in ['VENTA', 'RETIRO', 'SELL', 'OUT']):
                    m_venta = amt
                else:
                    m_compra = amt

            cot_buy_idx = col_map.get('cot_buy')
            cot_sell_idx = col_map.get('cot_sell')
            price_idx = col_map.get('price')

            c_compra = clean_decimal(row[cot_buy_idx]) if cot_buy_idx is not None and cot_buy_idx < len(row) else (clean_decimal(row[price_idx]) if price_idx is not None and price_idx < len(row) and m_compra > 0 else 0.0)
            c_venta = clean_decimal(row[cot_sell_idx]) if cot_sell_idx is not None and cot_sell_idx < len(row) else (clean_decimal(row[price_idx]) if price_idx is not None and price_idx < len(row) and m_venta > 0 else 0.0)

            fiat_idx = col_map.get('fiat')
            m_ars = clean_decimal(row[fiat_idx]) if fiat_idx is not None and fiat_idx < len(row) else 0.0

            if m_ars == 0.0:
                if m_compra > 0 and c_compra > 0:
                    m_ars = m_compra * c_compra
                elif m_venta > 0 and c_venta > 0:
                    m_ars = m_venta * c_venta

            notes_idx = col_map.get('notes')
            raw_notes = str(row[notes_idx]).strip() if notes_idx is not None and notes_idx < len(row) and row[notes_idx] is not None and str(row[notes_idx]).strip() not in ['', 'nan', 'None'] else ''

            tx_dict = create_transaction(
                dt, row_ex, tipo_op, raw_moneda,
                m_compra, m_venta, c_compra, c_venta, m_ars,
                comentario=raw_notes
            )

            records.append(tx_dict)
        except Exception as row_err:
            print(f"Error procesando fila {idx} en multi exchange excel: {row_err}")
            continue

    sample = records[:5]
    return records, sample
